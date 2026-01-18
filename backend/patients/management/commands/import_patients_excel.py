from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.dateparse import parse_date


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]+", "", s)
    return s


def _to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _to_date(v):
    if v is None or v == "":
        return None
    if hasattr(v, "date") and not isinstance(v, str):
        try:
            return v.date()
        except Exception:
            pass
    if isinstance(v, datetime):
        return v.date()
    s = _to_str(v)
    d = parse_date(s)
    return d


def _map_sexe(v: str | None):
    s = _norm(_to_str(v))
    if s in ("f", "femme"):
        return "F"
    if s in ("m", "homme"):
        return "M"
    return None


def _map_zone(v: str | None):
    s = _norm(_to_str(v))
    if s in ("grandbassam", "grand_bassam", "bassam"):
        return "GRAND_BASSAM"
    if s in ("bonoua",):
        return "BONOUA"
    return None


@dataclass
class ImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0


class Command(BaseCommand):
    help = "Import patients depuis un fichier Excel (.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("excel_path", help="Chemin vers le fichier Excel (.xlsx)")
        parser.add_argument("--sheet", default=None, help="Nom de feuille (par défaut: première feuille)")
        parser.add_argument("--dry-run", action="store_true", help="N'écrit rien en base, affiche seulement le résultat")
        parser.add_argument(
            "--update-existing",
            action="store_true",
            help="Met à jour les patients existants (matching code_patient). Par défaut: skip.",
        )
        parser.add_argument(
            "--map",
            action="append",
            default=[],
            help="Mapping colonnes Excel -> champs. Ex: --map nom=Nom --map prenoms=Prenoms",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise CommandError("openpyxl manquant. Installe: pip install -r requirements.txt") from exc

        from patients.models import Patient

        excel_path = options["excel_path"]
        sheet_name = options["sheet"]
        dry_run = bool(options["dry_run"])
        update_existing = bool(options["update_existing"])

        # Mapping par défaut: clés = champs Patient, valeurs = header normalisé attendu
        default_expected_headers = {
            "code_patient": ["code_patient", "code", "codepat", "id", "identifiant"],
            "nom": ["nom", "name"],
            "prenoms": ["prenoms", "prenom", "prenoms", "firstname"],
            "telephone": ["telephone", "tel", "phone", "mobile"],
            "adresse": ["adresse", "address"],
            "zone": ["zone", "localite", "commune"],
            "sexe": ["sexe", "genre", "sex"],
            "date_naissance": ["date_naissance", "datenaissance", "dob", "naissance"],
            "antecedents": ["antecedents", "antecedant", "antecedentsmedicaux"],
        }

        # Surcharges mapping via --map champ=ColonneExcel
        overrides: dict[str, str] = {}
        for item in options["map"]:
            if "=" not in item:
                raise CommandError("Format attendu pour --map: champ=ColonneExcel")
            k, v = item.split("=", 1)
            overrides[_norm(k)] = _norm(v)

        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise CommandError("Fichier Excel vide.")

        header_row = [
            _norm(_to_str(h))
            for h in (rows[0] or [])
        ]

        header_index = {h: idx for idx, h in enumerate(header_row) if h}

        # Résout index de colonne pour chaque champ
        field_to_col: dict[str, int] = {}
        for field, candidates in default_expected_headers.items():
            if field in overrides:
                h = overrides[field]
                if h in header_index:
                    field_to_col[field] = header_index[h]
                continue

            for cand in candidates:
                if cand in header_index:
                    field_to_col[field] = header_index[cand]
                    break

        required = ["nom", "prenoms"]
        missing_required = [f for f in required if f not in field_to_col]
        if missing_required:
            raise CommandError(
                "Colonnes obligatoires introuvables: " + ", ".join(missing_required) +
                ". Utilise --map pour préciser."
            )

        stats = ImportStats()

        def cell(field: str, row: tuple):
            idx = field_to_col.get(field)
            if idx is None:
                return None
            if idx >= len(row):
                return None
            return row[idx]

        def build_code_patient(nom: str, prenoms: str, row_num: int) -> str:
            base = _norm(nom)[:6] + _norm(prenoms)[:6]
            base = (base or "patient")[:12]
            return f"{base}_{row_num}"

        # Import sous transaction. En mode dry-run, on force un rollback.
        with transaction.atomic():
            for i, row in enumerate(rows[1:], start=2):
                try:
                    nom = _to_str(cell("nom", row))
                    prenoms = _to_str(cell("prenoms", row))
                    if not nom or not prenoms:
                        stats.skipped += 1
                        continue

                    code_patient = _to_str(cell("code_patient", row))
                    if not code_patient:
                        code_patient = build_code_patient(nom, prenoms, i)

                    sexe = _map_sexe(cell("sexe", row))
                    zone = _map_zone(cell("zone", row))
                    date_naissance = _to_date(cell("date_naissance", row))

                    defaults = {
                        "nom": nom,
                        "prenoms": prenoms,
                        "telephone": _to_str(cell("telephone", row)),
                        "adresse": _to_str(cell("adresse", row)),
                        "antecedents": _to_str(cell("antecedents", row)),
                    }
                    if sexe is not None:
                        defaults["sexe"] = sexe
                    if zone is not None:
                        defaults["zone"] = zone
                    if date_naissance is not None:
                        defaults["date_naissance"] = date_naissance

                    existing = Patient.objects.filter(code_patient=code_patient).first()
                    if existing:
                        if not update_existing:
                            stats.skipped += 1
                            continue
                        for k, v in defaults.items():
                            setattr(existing, k, v)
                        if not dry_run:
                            existing.save()
                        stats.updated += 1
                        continue

                    if not dry_run:
                        Patient.objects.create(code_patient=code_patient, **defaults)
                    stats.created += 1

                except Exception:
                    stats.errors += 1

            if dry_run:
                transaction.set_rollback(True)

        self.stdout.write(
            self.style.SUCCESS(
                f"Import terminé. created={stats.created} updated={stats.updated} skipped={stats.skipped} errors={stats.errors}"
            )
        )
